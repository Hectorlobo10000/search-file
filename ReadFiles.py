import os
from openpyxl import Workbook

file = raw_input('Nombre del archivo: ')
fileName = file + '.txt'

wb = Workbook()
dest_filename = file + '.xlsx'
ws = wb.active
ws.title = 'English'

openedFile = open(fileName, 'r+')
ws.cell(column=1, row=1, value='Carpeta')
ws.cell(column=2, row=1, value='Nombre de archivo')
ws.cell(column=3, row=1, value='Duracion')
ws.cell(column=4, row=1, value='Tamano en bytes')
ws.cell(column=5, row=1, value='Observacion')
row = 2
for line in openedFile:
  if not 'Total seconds' in line:
    tokens = line.split(' --- ')
    if not '\n' in tokens[0]:
      if 'ERROR DE ARCHIVO' in line:
        for col in range(1, 6):
          ws.cell(column=col, row=row, value=tokens[col - 1].replace('\n', ''))
      elif 'FILE MOVED' in line:
        for col in range(1, 6):
          ws.cell(column=col, row=row, value=tokens[col - 1].replace('\n', ''))
      else:
        for col in range(1, 5):
          ws.cell(column=col, row=row, value=tokens[col - 1].replace('\n', ''))
      row = row + 1
      print(row - 1)
  else: 
    break

openedFile.close()
wb.save(filename = dest_filename)